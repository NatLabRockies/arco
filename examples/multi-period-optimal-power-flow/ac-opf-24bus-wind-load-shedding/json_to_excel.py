#!/usr/bin/env python3
"""Convert arco result.json into the same Excel layout as results.xlsx.

This script reads an existing JSON solve output (no re-solve) and writes
three sheets compatible with the current exporter format:
- classic
- classic2
- classic3
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook

SBASE = 100.0


def _load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)
    if not isinstance(payload, dict):
        raise ValueError("result JSON root must be an object")
    return payload


def _load_generator_bus_map(generators_csv: Path) -> dict[str, int]:
    mapping: dict[str, int] = {}
    with generators_csv.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            mapping[str(row["generator"])] = int(row["bus"])
    return mapping


def _report_rows(summary: dict[str, Any], name: str) -> tuple[list[str], list[dict[str, Any]]]:
    for report in summary.get("reports", []):
        if report.get("name") == name:
            return list(report.get("index", [])), list(report.get("values", []))
    return [], []


def _row_value(row: dict[str, Any], index_columns: list[str]) -> float:
    for key, value in row.items():
        if key not in index_columns and isinstance(value, (int, float)):
            return float(value)
    if "value" in row and isinstance(row["value"], (int, float)):
        return float(row["value"])
    raise ValueError(f"report row has no numeric value column: {row}")


def _dual_lmp_map(
    summary: dict[str, Any],
    dual_name: str,
    sign_scale: float,
) -> dict[tuple[int, int], float]:
    lmp: dict[tuple[int, int], float] = {}
    for dual in summary.get("dual_reports", []):
        if dual.get("name") != dual_name:
            continue
        for item in dual.get("values", []):
            bus = item.get("i")
            t = item.get("t")
            value = item.get("value")
            if isinstance(bus, int) and isinstance(t, int) and isinstance(value, (int, float)):
                lmp[(bus, t)] = round(sign_scale * float(value) / SBASE, 4)
                continue

            instance = str(item.get("instance", ""))
            matched = re.search(r"\[\s*(\d+)\s*,\s*(\d+)\s*\]", instance)
            if matched and isinstance(value, (int, float)):
                lmp[(int(matched.group(1)), int(matched.group(2)))] = round(
                    sign_scale * float(value) / SBASE,
                    4,
                )
    return lmp


def convert_json_to_excel(result_json: Path, generators_csv: Path, output_xlsx: Path) -> None:
    summary = _load_json(result_json)
    gen_bus = _load_generator_bus_map(generators_csv)

    theta_index, theta_rows = _report_rows(summary, "va")
    if not theta_rows:
        theta_index, theta_rows = _report_rows(summary, "theta")
    v_index, v_rows = _report_rows(summary, "v")
    pg_index, pg_rows = _report_rows(summary, "pg")
    qg_index, qg_rows = _report_rows(summary, "qg")
    pw_index, pw_rows = _report_rows(summary, "pw")
    shed_index, shed_rows = _report_rows(summary, "lsh")
    if not shed_rows:
        shed_index, shed_rows = _report_rows(summary, "shed")
    # Match GAMS sign convention for nodal prices in this AC-OPF export.
    lmp_p = _dual_lmp_map(summary, "power_balance", -1.0)
    lmp_q = _dual_lmp_map(summary, "reactive_power_balance", 1.0)

    theta_map: dict[tuple[int, int], float] = {}
    for row in theta_rows:
        bus = int(row["i"])
        t = int(row["t"])
        theta_map[(bus, t)] = _row_value(row, theta_index)

    v_map: dict[tuple[int, int], float] = {}
    for row in v_rows:
        bus = int(row["i"])
        t = int(row["t"])
        v_map[(bus, t)] = _row_value(row, v_index)

    pw_map: dict[tuple[int, int], float] = {}
    for row in pw_rows:
        bus = int(row["i"])
        t = int(row["t"])
        pw_map[(bus, t)] = _row_value(row, pw_index) * SBASE

    shed_map: dict[tuple[int, int], float] = {}
    for row in shed_rows:
        bus = int(row["i"])
        t = int(row["t"])
        shed_map[(bus, t)] = _row_value(row, shed_index) * SBASE

    pg_bus_map: dict[tuple[int, int], float] = {}
    for row in pg_rows:
        generator = str(row["g"])
        bus = gen_bus[generator]
        t = int(row["t"])
        pg_bus_map[(bus, t)] = pg_bus_map.get((bus, t), 0.0) + _row_value(row, pg_index) * SBASE

    qg_bus_map: dict[tuple[int, int], float] = {}
    for row in qg_rows:
        generator = str(row["g"])
        bus = gen_bus[generator]
        t = int(row["t"])
        qg_bus_map[(bus, t)] = qg_bus_map.get((bus, t), 0.0) + _row_value(row, qg_index) * SBASE

    all_bus = sorted(
        {bus for bus, _ in theta_map.keys()}
        | {bus for bus, _ in v_map.keys()}
        | {bus for bus, _ in pw_map.keys()}
        | {bus for bus, _ in pg_bus_map.keys()}
        | {bus for bus, _ in qg_bus_map.keys()}
        | {bus for bus, _ in shed_map.keys()}
    )
    all_t = sorted(
        {t for _, t in theta_map.keys()}
        | {t for _, t in v_map.keys()}
        | {t for _, t in pw_map.keys()}
        | {t for _, t in pg_bus_map.keys()}
        | {t for _, t in qg_bus_map.keys()}
        | {t for _, t in shed_map.keys()}
    )

    workbook = Workbook()
    ws_classic = workbook.active
    ws_classic.title = "classic"
    ws_classic.append(["t", "bus", "V", "Angle", "Pg", "Qg", "Pw", "Shed", "LMP_P", "LMP_Q"])

    ws_classic2 = workbook.create_sheet("classic2")
    ws_classic3 = workbook.create_sheet("classic3")

    for t in all_t:
        for bus in all_bus:
            theta = theta_map.get((bus, t), 0.0)
            voltage = v_map.get((bus, t), 1.0)
            pg = pg_bus_map.get((bus, t), 0.0)
            qg = qg_bus_map.get((bus, t), 0.0)
            pw = pw_map.get((bus, t), 0.0)
            shed = shed_map.get((bus, t), 0.0)
            lmp_p_value = lmp_p.get((bus, t), 0.0)
            lmp_q_value = lmp_q.get((bus, t), 0.0)
            ws_classic.append([t, bus, voltage, theta, pg, qg, pw, shed, lmp_p_value, lmp_q_value])

    # Format voltage with 6 decimals; keep other columns as before.
    for row in ws_classic.iter_rows(min_row=2, max_row=ws_classic.max_row):
        row[2].number_format = "0.000000"
        row[3].number_format = "0.000000"
        row[7].number_format = "0.0000"
        row[8].number_format = "0.0000"
        row[9].number_format = "0.0000"

    generator_buses = sorted(set(gen_bus.values()))
    time_header = [""] + [f"t{tt}" for tt in all_t]
    ws_classic2.append(time_header)
    ws_classic3.append(time_header)

    for bus in generator_buses:
        ws_classic2.append([bus] + [pg_bus_map.get((bus, tt), 0.0) for tt in all_t])
        ws_classic3.append([bus] + [qg_bus_map.get((bus, tt), 0.0) for tt in all_t])

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_xlsx)


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert arco result.json to Excel workbook.")
    default_dir = Path(__file__).resolve().parent
    parser.add_argument(
        "--input-json",
        type=Path,
        default=default_dir / "result.json",
        help="Path to arco JSON output.",
    )
    parser.add_argument(
        "--generators-csv",
        type=Path,
        default=default_dir / "data" / "generators.csv",
        help="Path to generators.csv for mapping generator -> bus.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=default_dir / "results_from_json.xlsx",
        help="Output Excel workbook path.",
    )
    args = parser.parse_args()

    convert_json_to_excel(
        result_json=args.input_json,
        generators_csv=args.generators_csv,
        output_xlsx=args.output,
    )
    print(f"Wrote workbook: {args.output}")


if __name__ == "__main__":
    main()
