#!/usr/bin/env python3
"""Run the translated 24-bus case and export GAMS-style Excel outputs.

This exporter writes three sheets compatible with the original GAMS naming:
- classic  : bus-time report (V, Angle, Pg, Qg, LMP_P, LMP_Q + wind/shed fields)
- classic2 : bus-time active generation report
- classic3 : bus-time reactive generation report
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import subprocess
import sys
import threading
import time
from pathlib import Path
from typing import Any

from openpyxl import Workbook

SBASE = 100.0


def _load_generator_bus_map(generators_csv: Path) -> dict[str, int]:
    mapping: dict[str, int] = {}
    with generators_csv.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            mapping[str(row["generator"])] = int(row["bus"])
    return mapping


def _report_rows(summary: dict[str, Any], name: str) -> tuple[list[str], list[dict[str, Any]]]:
    for report in summary.get("reports", []):
        if report.get("name") == name:
            return list(report.get("index", [])), list(report.get("values", []))
    return [], []


def _row_value(row: dict[str, Any], index_columns: list[str]) -> float:
    for key, value in row.items():
        if key not in index_columns and isinstance(value, (int, float)):
            return float(value)
    if "value" in row and isinstance(row["value"], (int, float)):
        return float(row["value"])
    raise ValueError(f"report row has no numeric value column: {row}")


def _dual_lmp_map(
    summary: dict[str, Any],
    dual_name: str,
    sign_scale: float,
) -> dict[tuple[int, int], float]:
    lmp: dict[tuple[int, int], float] = {}
    for dual in summary.get("dual_reports", []):
        if dual.get("name") != dual_name:
            continue
        for item in dual.get("values", []):
            instance = str(item.get("instance", ""))
            value = float(item.get("value", 0.0))
            matched = re.search(r"\[\s*(\d+)\s*,\s*(\d+)\s*\]", instance)
            if not matched:
                continue
            bus = int(matched.group(1))
            t = int(matched.group(2))
            lmp[(bus, t)] = round(sign_scale * value / SBASE, 4)
    return lmp


def _run_arco(
    repo_root: Path,
    model_path: Path,
    *,
    timeout_seconds: int,
) -> dict[str, Any]:
    command = [
        "cargo",
        "run",
        "--release",
        "-q",
        "-p",
        "arco-cli",
        "--",
        "run",
        "--solver-log",
        str(model_path),
    ]

    command_with_ipopt = [
        "cargo",
        "run",
        "--release",
        "-p",
        "arco-cli",
        "--features",
        "ipopt",
        "--",
        "run",
        "--solver-log",
        str(model_path),
    ]

    def _parse_summary(stdout_text: str) -> dict[str, Any]:
        payload = stdout_text.strip()
        if not payload:
            raise RuntimeError("arco-cli produced no JSON summary on stdout")

        try:
            return json.loads(payload)
        except json.JSONDecodeError:
            pass

        # If extra stdout lines are present, parse the last JSON-like line.
        for line in reversed(payload.splitlines()):
            candidate = line.strip()
            if not candidate.startswith("{"):
                continue
            try:
                return json.loads(candidate)
            except json.JSONDecodeError:
                continue

        raise RuntimeError("unable to parse arco-cli JSON summary from stdout")

    def _run(command_to_run: list[str]) -> tuple[subprocess.CompletedProcess[str], float]:
        print(
            f"[export_results] running solver command (timeout={timeout_seconds}s): "
            + " ".join(command_to_run),
            file=sys.stderr,
        )
        started = time.time()

        process = subprocess.Popen(
            command_to_run,
            cwd=repo_root,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )

        stderr_lines: list[str] = []

        def _stream_stderr() -> None:
            assert process.stderr is not None
            for line in process.stderr:
                stderr_lines.append(line)
                # Stream solver progress/iterations live to terminal.
                print(line, end="", file=sys.stderr)

        stderr_thread = threading.Thread(target=_stream_stderr, daemon=True)
        stderr_thread.start()

        try:
            stdout_text, _ = process.communicate(timeout=timeout_seconds)
        except subprocess.TimeoutExpired as error:
            process.kill()
            stdout_text, _ = process.communicate()
            stderr_thread.join(timeout=1.0)
            raise subprocess.TimeoutExpired(
                cmd=error.cmd,
                timeout=error.timeout,
                output=stdout_text,
                stderr="".join(stderr_lines),
            ) from error

        stderr_thread.join(timeout=1.0)
        stderr_text = "".join(stderr_lines)

        completed = subprocess.CompletedProcess(
            args=command_to_run,
            returncode=process.returncode,
            stdout=stdout_text,
            stderr=stderr_text,
        )

        if completed.returncode != 0:
            raise subprocess.CalledProcessError(
                returncode=completed.returncode,
                cmd=command_to_run,
                output=completed.stdout,
                stderr=completed.stderr,
            )

        elapsed = time.time() - started
        print(
            f"[export_results] solver finished in {elapsed:.1f}s",
            file=sys.stderr,
        )
        return completed, elapsed

    try:
        completed, _elapsed = _run(command)
    except subprocess.TimeoutExpired as error:
        guidance = (
            "arco-cli solve exceeded timeout and was terminated.\n\n"
            f"Timeout: {timeout_seconds} seconds\n"
            "Model: "
            f"{model_path}\n\n"
            "Try one of:\n"
            "1) Increase exporter timeout, e.g. --timeout 1800\n"
            "2) Run solver directly to inspect progress:\n"
            f"   cargo run -p arco-cli --features ipopt -- run {model_path}\n"
            "3) Relax/tighten model or set a solver time limit in your workflow."
        )
        raise RuntimeError(guidance) from error
    except subprocess.CalledProcessError as error:
        stderr = (error.stderr or "").strip()
        if (
            "IPOPT solver backend is not available in this build" in stderr
            or "arco::driver::backend_not_available" in stderr
        ):
            print(
                "[export_results] retrying with IPOPT-enabled build; initial compile may take a while before iterations appear",
                file=sys.stderr,
            )
            try:
                completed, _elapsed = _run(command_with_ipopt)
            except subprocess.TimeoutExpired as retry_timeout:
                guidance = (
                    "arco-cli solve exceeded timeout and was terminated.\n\n"
                    f"Timeout: {timeout_seconds} seconds\n"
                    "Model: "
                    f"{model_path}\n\n"
                    "Try one of:\n"
                    "1) Increase exporter timeout, e.g. --timeout 1800\n"
                    "2) Run solver directly to inspect progress:\n"
                    f"   cargo run -p arco-cli --features ipopt -- run --solver-log {model_path}\n"
                    "3) Relax/tighten model or set a solver time limit in your workflow."
                )
                raise RuntimeError(guidance) from retry_timeout
            except subprocess.CalledProcessError as retry_error:
                retry_stderr = (retry_error.stderr or "").strip()
                if (
                    "failed to run custom build command for `ipopt-sys`" in retry_stderr
                    or "Ipopt_INCLUDE_DIR" in retry_stderr
                ):
                    guidance = (
                        "IPOPT backend was requested, and exporter retried with `--features ipopt`, "
                        "but local IPOPT headers/libs were not found by ipopt-sys.\n\n"
                        "Try:\n"
                        "1) Install/verify IPOPT development files (Homebrew ipopt on macOS).\n"
                        "2) Ensure include dir is discoverable (for this machine: /opt/homebrew/Cellar/ipopt/3.14.17/include/coin-or).\n"
                        "3) Re-run exporter.\n\n"
                        "If you prefer installing the binary first in this workspace, use:\n"
                        "   cargo install --path crates/arco-cli --features ipopt"
                    )
                    raise RuntimeError(guidance) from retry_error

                diagnostic = (
                    retry_stderr or (retry_error.stdout or "").strip() or str(retry_error)
                )
                raise RuntimeError(
                    "arco-cli run failed after retrying with IPOPT feature:\n"
                    f"{diagnostic}"
                ) from retry_error

        if "does not support nonlinear algebra" in stderr:
            guidance = (
                "arco-cli is currently using a linear backend (HiGHS/Xpress), "
                "but this AC-OPF model is nonlinear and requires IPOPT.\n\n"
                "1) Build arco-cli with IPOPT feature:\n"
                "   cargo install --path crates/arco-cli --features ipopt\n"
                "2) Select IPOPT backend:\n"
                "   cargo run -q -p arco-cli -- solver set ipopt\n"
                "3) Re-run this exporter.\n\n"
                "If IPOPT build fails, ensure IPOPT headers/libs are discoverable on your machine."
            )
            raise RuntimeError(guidance) from error

        diagnostic = stderr or (error.stdout or "").strip() or str(error)
        raise RuntimeError(f"arco-cli run failed:\n{diagnostic}") from error

    return _parse_summary(completed.stdout)


def export_excel(
    repo_root: Path,
    model_path: Path,
    output_xlsx: Path,
    *,
    timeout_seconds: int,
) -> None:
    print("[export_results] starting export", file=sys.stderr)
    summary = _run_arco(
        repo_root=repo_root,
        model_path=model_path,
        timeout_seconds=timeout_seconds,
    )

    gen_bus = _load_generator_bus_map(model_path.parent / "data" / "generators.csv")

    theta_index, theta_rows = _report_rows(summary, "theta")
    v_index, v_rows = _report_rows(summary, "v")
    pg_index, pg_rows = _report_rows(summary, "pg")
    qg_index, qg_rows = _report_rows(summary, "qg")
    pw_index, pw_rows = _report_rows(summary, "pw")
    shed_index, shed_rows = _report_rows(summary, "lsh")
    if not shed_rows:
        shed_index, shed_rows = _report_rows(summary, "shed")
    # Match GAMS sign convention for nodal prices in this AC-OPF export.
    lmp_p = _dual_lmp_map(summary, "power_balance", -1.0)
    lmp_q = _dual_lmp_map(summary, "reactive_power_balance", 1.0)

    theta_map: dict[tuple[int, int], float] = {}
    for row in theta_rows:
        bus = int(row["i"])
        t = int(row["t"])
        theta_map[(bus, t)] = _row_value(row, theta_index)

    v_map: dict[tuple[int, int], float] = {}
    for row in v_rows:
        bus = int(row["i"])
        t = int(row["t"])
        v_map[(bus, t)] = _row_value(row, v_index)

    pw_map: dict[tuple[int, int], float] = {}
    for row in pw_rows:
        bus = int(row["i"])
        t = int(row["t"])
        pw_map[(bus, t)] = _row_value(row, pw_index) * SBASE

    shed_map: dict[tuple[int, int], float] = {}
    for row in shed_rows:
        bus = int(row["i"])
        t = int(row["t"])
        shed_map[(bus, t)] = _row_value(row, shed_index) * SBASE

    pg_bus_map: dict[tuple[int, int], float] = {}
    for row in pg_rows:
        generator = str(row["g"])
        bus = gen_bus[generator]
        t = int(row["t"])
        pg_bus_map[(bus, t)] = pg_bus_map.get((bus, t), 0.0) + _row_value(row, pg_index) * SBASE

    qg_bus_map: dict[tuple[int, int], float] = {}
    for row in qg_rows:
        generator = str(row["g"])
        bus = gen_bus[generator]
        t = int(row["t"])
        qg_bus_map[(bus, t)] = qg_bus_map.get((bus, t), 0.0) + _row_value(row, qg_index) * SBASE

    all_bus = sorted(
        {bus for bus, _ in theta_map.keys()}
        | {bus for bus, _ in v_map.keys()}
        | {bus for bus, _ in pw_map.keys()}
        | {bus for bus, _ in pg_bus_map.keys()}
        | {bus for bus, _ in qg_bus_map.keys()}
        | {bus for bus, _ in shed_map.keys()}
    )
    all_t = sorted(
        {t for _, t in theta_map.keys()}
        | {t for _, t in v_map.keys()}
        | {t for _, t in pw_map.keys()}
        | {t for _, t in pg_bus_map.keys()}
        | {t for _, t in qg_bus_map.keys()}
        | {t for _, t in shed_map.keys()}
    )

    workbook = Workbook()
    ws_classic = workbook.active
    ws_classic.title = "classic"
    ws_classic.append(["t", "bus", "V", "Angle", "Pg", "Qg", "Pw", "Shed", "LMP_P", "LMP_Q"])

    ws_classic2 = workbook.create_sheet("classic2")
    ws_classic3 = workbook.create_sheet("classic3")

    for t in all_t:
        for bus in all_bus:
            theta = theta_map.get((bus, t), 0.0)
            voltage = v_map.get((bus, t), 1.0)
            pg = pg_bus_map.get((bus, t), 0.0)
            qg = qg_bus_map.get((bus, t), 0.0)
            pw = pw_map.get((bus, t), 0.0)
            shed = shed_map.get((bus, t), 0.0)
            lmp_p_value = lmp_p.get((bus, t), 0.0)
            lmp_q_value = lmp_q.get((bus, t), 0.0)

            ws_classic.append([t, bus, voltage, theta, pg, qg, pw, shed, lmp_p_value, lmp_q_value])

    # Format voltage with 6 decimals; keep other columns as before.
    for row in ws_classic.iter_rows(min_row=2, max_row=ws_classic.max_row):
        row[2].number_format = "0.000000"
        row[3].number_format = "0.000000"
        row[7].number_format = "0.0000"
        row[8].number_format = "0.0000"
        row[9].number_format = "0.0000"

    generator_buses = sorted(set(gen_bus.values()))
    time_header = [""] + [f"t{tt}" for tt in all_t]
    ws_classic2.append(time_header)
    ws_classic3.append(time_header)

    for bus in generator_buses:
        ws_classic2.append([bus] + [pg_bus_map.get((bus, tt), 0.0) for tt in all_t])
        ws_classic3.append([bus] + [qg_bus_map.get((bus, tt), 0.0) for tt in all_t])

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_xlsx)
    print(f"[export_results] wrote workbook: {output_xlsx}", file=sys.stderr)


def main() -> None:
    parser = argparse.ArgumentParser(description="Export translated 24-bus OPF results to Excel.")
    parser.add_argument(
        "--repo-root",
        type=Path,
        default=Path(__file__).resolve().parents[3],
        help="Repository root containing Cargo.toml.",
    )
    parser.add_argument(
        "--model",
        type=Path,
        default=Path(__file__).resolve().parent / "input.kdl",
        help="Path to the KDL model file.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(__file__).resolve().parent / "results.xlsx",
        help="Output Excel workbook path.",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=900,
        help="Maximum seconds to wait for arco-cli solve before aborting.",
    )
    args = parser.parse_args()

    try:
        export_excel(
            repo_root=args.repo_root,
            model_path=args.model,
            output_xlsx=args.output,
            timeout_seconds=args.timeout,
        )
    except RuntimeError as error:
        print(str(error), file=sys.stderr)
        raise SystemExit(1) from error


if __name__ == "__main__":
    main()
